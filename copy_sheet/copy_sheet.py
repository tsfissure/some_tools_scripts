#coding=utf-8
'''
Created on 2018/5/5 17:00:41

@author: tys
'''

import os
from win32com.client import Dispatch as dp
import openpyxl

class CopySheet(object):

    def __init__(self):
        self.path = ""
        self.mainfile = ""
        self.xl = None
        self.mainBookXL = None
        self.bookXL = None
        self.logFile = open("log.log", "w")


    def loadConfig(self):
        self.logMsg(u'loading config...')
        cfg = open("config", "r")
        for line in cfg:
            lst = line[:-1].split('=')
            if 2 == len(lst):
                lst[0] = lst[0].strip()
                lst[1] = lst[1].strip()
                if "path" == lst[0]:
                    self.path = lst[1]
                elif "mainfile" == lst[0]:
                    self.mainfile = lst[1]
        cfg.close()
    
    def listSubFiles(self):
        lst = os.listdir(self.path)
        rlt = []
        for fd in lst:
            if os.path.isfile(os.path.join(self.path, fd)) and fd.endswith('.xlsx'):
                rlt.append(os.path.join(self.path, fd))
        return rlt

    def backupMainFile(self):
        absName = os.path.join(self.path, self.mainfile)
        self.logMsg(u'back up file[%s]' % absName)
        if os.path.isfile(absName):
            bakName = os.path.join(self.path, self.mainfile + ".bak")
            if os.path.isfile(bakName):
                os.remove(bakName)
            open(bakName, "wb").write(open(absName, "rb").read())

    def prepare(self):
        self.xl = dp("Excel.Application")

    def closeMainXL(self):
        if self.mainBookXL:
            self.mainBookXL.Close()
            self.mainBookXL = None

    def closeBookXL(self):
        if self.bookXL:
            self.bookXL.Close(SaveChanges = True)
            self.bookXL = None

    def logMsg(self, msg):
        print msg
        # TODO log to file


    def tryCopy(self, mainNames, fd):
        if fd.endswith(self.mainfile): return
        self.logMsg(u'sync [%s] start...' % fd)
        book = openpyxl.load_workbook(fd) # 用openpyxl打开excel 方便删除
        names = book.sheetnames[:]
        inter = [i for i in xrange(len(names)) if names[i] in mainNames]
        if len(inter) > 0:

            if inter[-1] == len(names) - 1: # 全部都需要同步或者最后一个需要同步
                book.create_sheet(title = 'tmpsheet')

            for i in inter:
                book.remove(book[names[i]])
            book.save(fd)
            book.close()

            self.bookXL = self.xl.Workbooks.Open(Filename = fd)
            bookNames = [self.bookXL.Worksheets(i + 1).Name for i in xrange(self.bookXL.Worksheets.Count)]
            for i in inter:
                self.logMsg(u'sync [%s]->sheet[%s]...' % (fd, names[i]))
                mainSheet = self.mainBookXL.Worksheets(names[i])
                mainSheet.Copy(Before = self.bookXL.Worksheets(i + 1))
            self.closeBookXL()

            if inter[-1] == len(names) - 1: # 全部都需要同步或者最后一个需要同步
                book = openpyxl.load_workbook(fd) # 需要重新打开删除
                self.logMsg(book.sheetnames)
                book.remove(book['tmpsheet'])
                book.save(fd)
        self.logMsg(u'sync [%s] end...' % fd)

    def gao(self):
        self.logMsg(u'sync start...')
        absName = os.path.join(self.path, self.mainfile)
        tmpName = os.path.join(self.path, self.mainfile + ".tmp.xlsx")
        files = self.listSubFiles()
        self.mainBookXL = self.xl.Workbooks.Open(Filename=absName)
        mainNames = [self.mainBookXL.Worksheets(i + 1).Name for i in xrange(self.mainBookXL.Worksheets.Count)]
        for fd in files:
            self.tryCopy(mainNames, fd)
        self.mainBookXL.Close()
        self.xl.Quit()
        self.logMsg(u'sync end...')

    def exceptQuit(self):
        if self.mainBookXL:
            self.mainBookXL.Close()
        if self.bookXL:
            self.bookXL.Close(SaveChanges = True)
        if self.xl:
            self.xl.Quit()
            del self.xl

    def trygao(self):
        try:
            self.backupMainFile()
            self.prepare()
            self.gao()
        except Exception, e:
            try:
                self.exceptQuit()
            except:
                pass
            print('error:', unicode(e))


instance = CopySheet()

if __name__ == '__main__':
    instance.loadConfig()
    instance.trygao()
    os.system("pause")

